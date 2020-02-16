# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import io
import logging

# The followong limits are for the formatted Excel sheet
MAX_NR_MEETINGS = (24, 36)
MAX_NR_PERSONS = (36, 48)
TEMPLATES = ("templates/narvarokort_sthlm.xlsx",
             "templates/narvarokort_sthlm_big.xlsx")


def format_postnummer(postnummer):
    return postnummer.replace(" ", "")


class ExcelReportSthlm:
    def __init__(self, dak, semester):
        self.dak = dak
        self.semester = semester

    def getFilledInExcelSpreadsheet(self):
        deltagareCount = len(self.dak.kort.deltagare)
        ledareCount = len(self.dak.kort.ledare)
        totalPersons = deltagareCount + ledareCount
        eventCount = len(self.dak.kort.sammankomster)
        logging.info("%d personer och %d sammankomster" % (totalPersons, eventCount))
        if totalPersons <= MAX_NR_PERSONS[0] and eventCount <= MAX_NR_MEETINGS[0]:
            workbook = load_workbook(TEMPLATES[0])
        elif totalPersons <= MAX_NR_PERSONS[1] and eventCount <= MAX_NR_MEETINGS[1]:
            workbook = load_workbook(TEMPLATES[1])
        else:
            raise ValueError("För många personer %d eller sammankomster %d" % (totalPersons, eventCount))

        startRowPersons = 15
        ws = workbook.worksheets[0]
        ws['A1'] = "Närvarokort Nr %s" % self.dak.kort.naervarokort_nummer
        ws['AJ1'] = self.semester.year
        ws['A3'] = self.dak.foerenings_namn
        ws['A5'] = "Scouting"
        ws['C5'] = self.dak.kort.namn_paa_kort
        ws['A7'] = self.dak.kort.lokal
        if self.semester.ht:
            ws['D1'] = "HT-%d" % (self.semester.year % 100)
        else:
            ws['D1'] = "VT-%d" % (self.semester.year % 100)




        for index, sammankomst in enumerate(self.dak.kort.sammankomster):
            aktivitetColumn = 7 + index
            ws.cell(row=4, column=aktivitetColumn).value = sammankomst.aktivitet
            ws.cell(row=10, column=aktivitetColumn).value = sammankomst.get_start_time_string('%H')
            ws.cell(row=11, column=aktivitetColumn).value = sammankomst.get_stop_time_string('%H')

            ws.cell(row=12, column=aktivitetColumn).value = sammankomst.get_date_string('%m')
            ws.cell(row=13, column=aktivitetColumn).value = sammankomst.get_date_string('%d')
            for deltagareindex, deltagaren in enumerate(self.dak.kort.deltagare):
                if deltagaren in sammankomst.deltagare:
                    ws.cell(row=startRowPersons+deltagareindex, column=aktivitetColumn).value = "x"

            for ledarindex, ledaren in enumerate(self.dak.kort.ledare):
                if ledaren in sammankomst.ledare:
                    ws.cell(row=startRowPersons+ledarindex+deltagareCount, column=aktivitetColumn).value = "x"

        for index, deltagaren in enumerate(self.dak.kort.deltagare):
            deltagarenRow = startRowPersons+index
            ws.cell(row=deltagarenRow, column=2).value = deltagaren.foernamn
            ws.cell(row=deltagarenRow, column=3).value = deltagaren.efternamn
            ws.cell(row=deltagarenRow, column=4).value = 'k' if deltagaren.is_female() else 'm'
            ws.cell(row=deltagarenRow, column=5).value = format_postnummer(deltagaren.postnummer)
            ws.cell(row=deltagarenRow, column=6).value = deltagaren.personnummer[:4]

        for index, ledaren in enumerate(self.dak.kort.ledare):
            ledarenRow = startRowPersons+index+deltagareCount
            ws.cell(row=ledarenRow, column=2).value = ledaren.foernamn
            ws.cell(row=ledarenRow, column=3).value = ledaren.efternamn
            ws.cell(row=ledarenRow, column=4).value = 'k' if ledaren.is_female() else 'm'
            ws.cell(row=ledarenRow, column=5).value = format_postnummer(deltagaren.postnummer)
            ws.cell(row=ledarenRow, column=6).value = ledaren.personnummer[:4]

        bytesIO = io.BytesIO()
        workbook.save(bytesIO)
        return bytesIO.getvalue()
