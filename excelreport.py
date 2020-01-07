# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import io


class ExcelReport:
    def __init__(self, dak, semester):
        self.dak = dak
        self.semester = semester

    def getFilledInExcelSpreadsheet(self):
        workbook = load_workbook("templates/narvarokort.xlsx")
        numPersonRows = 64
        startRowPersons = 13
        ws = workbook.worksheets[0]
        ws['E1'] = self.dak.kort.naervarokort_nummer
        ws['I1'] = self.semester.year
        ws['D2'] = self.dak.kort.namn_paa_kort
        ws['D3'] = "Scouting"
        ws['D4'] = self.dak.kort.lokal
        if self.semester.ht:
            ws['C7'] = 'X'
        else:
            ws['C6'] = 'X'

        deltagareCount = len(self.dak.kort.deltagare)
        ledareCount = len(self.dak.kort.ledare)
        totalPersons = deltagareCount + ledareCount

        for index, sammankomst in enumerate(self.dak.kort.sammankomster):
            aktivitetColumn = 11+index
            ws.cell(row=2, column=aktivitetColumn).value = sammankomst.aktivitet
            ws.cell(row=7, column=aktivitetColumn).value = sammankomst.get_start_time_string('%H')
            ws.cell(row=9, column=aktivitetColumn).value = sammankomst.get_stop_time_string('%H')
            ws.cell(row=10, column=aktivitetColumn).value = sammankomst.get_date_string('%m')
            ws.cell(row=11, column=aktivitetColumn).value = sammankomst.get_date_string('%d')
            for deltagareindex, deltagaren in enumerate(self.dak.kort.deltagare):
                if deltagaren in sammankomst.deltagare:
                    ws.cell(row=startRowPersons+deltagareindex, column=aktivitetColumn).value = 1

            for ledarindex, ledaren in enumerate(self.dak.kort.ledare):
                if ledaren in sammankomst.ledare:
                    ws.cell(row=startRowPersons+ledarindex+deltagareCount, column=aktivitetColumn).value = 1

        for index, deltagaren in enumerate(self.dak.kort.deltagare):
            deltagarenRow = startRowPersons+index
            ws.cell(row=deltagarenRow, column=2).value = deltagaren.foernamn + " " + deltagaren.efternamn
            ws.cell(row=deltagarenRow, column=8).value = 'K' if deltagaren.is_female() else 'M'
            ws.cell(row=deltagarenRow, column=9).value = deltagaren.postnummer
            ws.cell(row=deltagarenRow, column=10).value = deltagaren.personnummer[0:8]

        for index, ledaren in enumerate(self.dak.kort.ledare):
            ledarenRow = startRowPersons+index+deltagareCount
            ws.cell(row=ledarenRow, column=2).value = "Ledare:"
            ws.cell(row=ledarenRow, column=3).value = ledaren.foernamn + " " + ledaren.efternamn
            ws.cell(row=ledarenRow, column=8).value = 'K' if ledaren.is_female() else 'M'
            ws.cell(row=ledarenRow, column=9).value = ledaren.postnummer
            ws.cell(row=ledarenRow, column=10).value = ledaren.personnummer[0:8]

        bytesIO = io.BytesIO()
        workbook.save(bytesIO)
        return bytesIO.getvalue()
